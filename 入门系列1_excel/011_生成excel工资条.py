from openpyxl import load_workbook, Workbook


def create_excel():
    wb = load_workbook('工资条.xlsx')
    sh = wb.active
    # 按照行数来获取
    title = ['工号', '姓名', '部门', '基本工资', '提成', '加班工资', '社保扣除', '考勤扣除', '应发工资', '邮箱']
    for i, row in enumerate(sh.rows):  # 通过枚举的形式拿到数据，i代表索引（从0开始），row代表sh中的每一行数据
        if i == 0:
            continue
        else:
            tmp_wb = Workbook()
            tmp_sheet = tmp_wb.active
            tmp_sheet.append(title)
            # 这里要取出每一行单元格中的数据
            # 第一种方法
            # row_value = []
            # for cell in row:
            #     row_value.append(cell.value)
            # 第二种方法
            row_value = [cell.value for cell in row]
            tmp_sheet.append(row_value)
            tmp_wb.save(
                f'./create_data/{row_value[1]}.xlsx')  # 每次读取一行数据以后，就要保存到一个新的excel工资表里面，注意，这里保存数据前，一个先创建出数据保存的文件夹


if __name__ == '__main__':
    create_excel()
