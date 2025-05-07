def create_new():
    from openpyxl import Workbook
    wb = Workbook()  # 此时默认是有一个Sheet1的工作簿的
    sh1 = wb.active
    sh2 = wb.create_sheet('数据')
    sh2 = wb.create_sheet('人员', 0)
    sh2 = wb.create_sheet('照汗青', 1)  # 不会占用之前的工作簿，之前的工作簿会顺序向后排列
    wb.save('08_通过openpyxl创建的excel表格.xlsx')


def set_value():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, colors
    bold_italic_30_font = Font(name='微软雅黑', size=30, italic=True, bold=True, color=colors.BLUE)
    bold_italic_20_font = Font(name='微软雅黑', size=20, italic=True, bold=True, color='9400D3')
    wb = Workbook()  # 此时默认是有一个Sheet1的工作簿的
    sh1 = wb.active
    sh1['A1'] = '肖顺康'
    sh1['A2'] = '肖顺康'
    sh1['A2'].font = bold_italic_30_font
    sh1['A3'] = '肖顺康'
    sh1['A3'].font = bold_italic_20_font
    wb.save('08_通过openpyxl创建的excel表格.xlsx')


def set_value2():
    from openpyxl import Workbook
    wb = Workbook()
    sh = wb.active
    data = ['python', 'c++', 'java']
    for i, d in enumerate(data):
        sh.cell(i + 1, 1).value = d  # 索引从1开始，所以要加上1
    wb.save('08_添加数据.xlsx')


def set_style2():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, colors
    wb = Workbook()
    sh = wb.active
    data = ['python', 'c++', 'java']
    sh.row_dimensions[1].height = 60  # 设置A1单元格的高度
    sh.column_dimensions['A'].width = 30  # 设置A1单元格的宽度
    for i, d in enumerate(data):
        sh.cell(i + 1, 1).value = d
        sh.cell(i + 1, 1).alignment = Alignment(horizontal='center', vertical='center')
    wb.save('08_添加数据.xlsx')


def set_merge():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, colors
    wb = Workbook()
    sh = wb.active
    sh.merge_cells('a1:c3')  # 合并单元格
    sh.merge_cells('d1:E5')
    sh['a1'] = '肖顺康'  # 合并单元格以后，写入数据的时候，只能从左上角的单元格开始写
    sh['a1'].alignment = Alignment(horizontal='center', vertical='center')  # 只要找到对应的单元格以后，就可以设置其样式
    sh['d1'] = 'xiao_shun_kang'
    sh['d1'].alignment = Alignment(horizontal='center', vertical='center')
    wb.save('08_添加数据.xlsx')


def set_img():
    from openpyxl import Workbook
    from datetime import date
    wb = Workbook()
    sh = wb.active
    rows = [
        ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
        [date(2020, 12, 1), 40, 30, 25],
        [date(2020, 12, 2), 40, 25, 30],
        [date(2020, 12, 3), 50, 30, 45],
        [date(2020, 12, 4), 30, 25, 40],
        [date(2020, 12, 5), 25, 35, 30],
        [date(2020, 12, 6), 20, 40, 35]
    ]
    sh.column_dimensions['A'].width = 30
    for row in rows:
        sh.append(row)  # 当数据是这种列表的形式的时候，可以通过这种方式批量添加数据
    from openpyxl.chart import LineChart, Reference, RadarChart  # 引入图表
    c1 = LineChart()  # 创建一个图表的对象
    c1.title = 'Line chart'
    c1.x_axis.title = 'x_axis'
    c1.y_axis.title = 'y_axis'
    data = Reference(sh, min_col=2, min_row=1, max_col=4, max_row=7)  # 选择依赖的数据,这里代表选择第一行到最后一行，第二列到最后一列的数据
    labels = Reference(sh, min_col=1, min_row=2, max_row=7)  # 设置图例
    c1.add_data(data, titles_from_data=True)  # 把数据添加到图表中
    c1.set_categories(labels)
    sh.add_chart(c1, 'a9')  # 把图标添加到工作簿中
    wb.save('08_添加数据.xlsx')


def set_img2():
    from openpyxl import Workbook
    from datetime import date
    wb = Workbook()
    sh = wb.active
    rows = [
        ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
        [date(2020, 12, 1), 40, 30, 25],
        [date(2020, 12, 2), 40, 25, 30],
        [date(2020, 12, 3), 50, 30, 45],
        [date(2020, 12, 4), 30, 25, 40],
        [date(2020, 12, 5), 25, 35, 30],
        [date(2020, 12, 6), 20, 40, 35]
    ]
    sh.column_dimensions['A'].width = 30
    for row in rows:
        sh.append(row)  # 当数据是这种列表的形式的时候，可以通过这种方式批量添加数据
    from openpyxl.chart import Reference, RadarChart  # 引入图表
    c1 = RadarChart()  # 创建一个图表的对象
    c1.type = 'filled'  # 设置图像的类型
    c1.title = '雷达图'
    c1.x_axis.title = 'x_axis'
    c1.y_axis.title = 'y_axis'
    data = Reference(sh, min_col=2, min_row=1, max_col=4, max_row=7)  # 选择依赖的数据,这里代表选择第一行到最后一行，第二列到最后一列的数据
    labels = Reference(sh, min_col=1, min_row=2, max_row=7)  # 设置类别的范围
    c1.add_data(data, titles_from_data=True)  # 把数据添加到图表中
    c1.set_categories(labels)  # 设置图例
    sh.add_chart(c1, 'a9')  # 把图标添加到工作簿中
    wb.save('08_添加数据.xlsx')


if __name__ == '__main__':
    # create_new()
    # set_value()
    # set_value2()
    # set_style2()
    # set_merge()
    set_img()
