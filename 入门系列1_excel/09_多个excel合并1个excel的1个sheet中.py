from openpyxl import load_workbook, Workbook
import os


def copy_data():
    wb = Workbook()
    sh = wb.active
    all_data = []  # 创建一个列表用来保存数据
    for name in os.listdir('data'):
        path = f'data/{name}'
        tmp_wb = load_workbook(path)
        tmp_sh = tmp_wb.active
        for r in range(1, tmp_sh.max_row+1):
            # 获得整行的数据
            row_value = []
            for c in range(1, tmp_sh.max_column+1):
                value = tmp_sh.cell(r, c).value
                row_value.append(value)  # 往列表里面添加读取到的数据
            # 把整行数据加入到全局数据里面
            if row_value not in all_data:
                all_data.append(row_value)
    for data in all_data:
        sh.append(data)  # 获得的data是一个列表，sh.apped(data)可以直接将列表里面的数据写入到工作簿中
    # 保存数据
    wb.save('09_合并的数据.xlsx')


if __name__ == '__main__':
    print(os.listdir('data'))  # 返回到上一级去拿取文件
    copy_data()
    pass
