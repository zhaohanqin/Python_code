from openpyxl import load_workbook, Workbook
from datetime import date
from openpyxl.styles import PatternFill


def create_excel():
    wb = Workbook()
    sh = wb.active
    rows = [
        ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
        [date(2020, 12, 1), 40, 30, 25],
        [date(2020, 12, 2), 40, 25, 30],
        [date(2020, 12, 3), 50, 30, 45],
        [date(2020, 12, 4), 30, 25, 25],
        [date(2020, 12, 6), 20, 40, 35],
    ]
    for row in rows:
        sh.append(row)
    # 修改样式，填充背景色

    # 设置对应的背景的颜色
    bak_color1 = PatternFill('solid', fgColor='BBFFFF')
    bak_color2 = PatternFill('solid', fgColor='C1FFC1')
    for r in range(1, sh.max_row + 1):
        if r % 2 == 0:
            for c in range(1, sh.max_column + 1):
                sh.cell(r, c).fill = bak_color1
        else:
            for c in range(1, sh.max_column + 1):
                sh.cell(r, c).fill = bak_color2

    wb.save(f'./create_data/012_隔行变色.xlsx')


if __name__ == '__main__':
    create_excel()
