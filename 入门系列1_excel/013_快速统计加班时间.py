from openpyxl import load_workbook, Workbook
from datetime import date
from openpyxl.styles import PatternFill


def create_excel():
    wb = Workbook()
    sh = wb.active
    rows = [
        ['Date', '姓名', '打卡时间'],
        [date(2020, 12, 1), '001', '18:50'],
        [date(2020, 12, 2), '002', '18:10'],
        [date(2020, 12, 3), '001', '18:02'],
        [date(2020, 12, 4), '001', '18:50'],
        [date(2020, 12, 6), '001', '20:50'],
        [date(2020, 12, 6), '006', '22:50'],
    ]
    for row in rows:
        sh.append(row)  # 将数据添加到工作簿里面
    wb.save('./create_data/013_打卡时间.xlsx')


# 读取数据
# 注意，这里创建新的工作簿的时候，是先将数据读取出来，然后再写入新的工作簿并创建工作簿中，xlutis的copy可能会简单一点
def statistics():
    wb = load_workbook('create_data/013_打卡时间.xlsx')
    sh = wb.active
    data = []  # 用来存储最后的工作簿
    for i in range(2, sh.max_row + 1):
        t_data = []  # 用来保存每一行的数据
        for j in range(1, sh.max_column + 1):
            t_data.append(sh.cell(i, j).value)  # 此时是一个列表
        # 对加班的时间进行统计,这里进行统计的操作一个在遍历完一行的数据后进行
        h, m = t_data[2].split(':')
        full_time = int(h) * 60 + int(m)
        tmp = full_time - 18 * 60
        t_data.append(str(tmp) + '分钟')  # 加在每一行列表数据的最后一列
        # 处理时间问题
        t_data[0] = t_data[0].date()  # 避免得到的时间后面有小时，分和秒
        data.append(t_data)
    # 保存数据
    wb = Workbook()
    sh = wb.active
    title = ['日期', '姓名', '打卡时间', '加班时间']
    sh.append(title)
    for d in data:
        sh.append(d)  # 按每一行的数据来将据写入新的工作簿中
    sh.column_dimensions['A'].width = 20  # 设置"A"列单元格的宽度
    sh.row_dimensions[1].height = 60  # 设置第1行单元格的高度
    wb.save('./create_data/013_打卡时间2.xlsx')  # 将数据保存


if __name__ == '__main__':
    create_excel()
    statistics()
