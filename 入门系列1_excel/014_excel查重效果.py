from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def read_excel():
    wb = load_workbook('create_data/013_打卡时间.xlsx')
    sh = wb.active
    index = []  # 设置一个列表，用来存储重复的数据，注意，这里存储的是重复数据的行数的索引（从0开始的索引）
    tmp = []  # 用来存储没有重复的数据
    # 这里用sh['B']的原因是根据打卡表里面的姓名来判断是否是重复数据，如果要根据其它列来判断，选择不同的列就行。sh['B']里面的数据的格式为:
    # (<Cell 'Sheet'.B1>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.B3>,
    # <Cell 'Sheet'.B4>, <Cell 'Sheet'.B5>, <Cell 'Sheet'.B6>, <Cell 'Sheet'.B7>)
    # 是元组的形式，拿到的数据为每个单元格（不是单元格对应的value），加上枚举后可以得到对应的行数的索引
    for i, c in enumerate(sh['B']):  # 这里用枚举的形式来拿第“B”列的数据，c为B列数据对应的单元格，i为数据对应的索引，行数
        if c.value not in tmp:  # 当不是重复数据的时候
            tmp.append(c.value)
        else:  # 由于索引默认是从0开始的，所以要进行加1的操作，openpyxl拿到的数据默认是从1开始的
            index.append(i + 1)  # 是重复数据的情况下，拿到对应的行的索引，方便后续对整行的操作
    for i, r in enumerate(sh.rows):  # 这里的r代表拿到每一行的所有的单元格
        if i + 1 in index:
            for c in r:  # 遍历每一行所有的单元格
                c.fill = PatternFill('solid', fgColor='BBFFFF')  # 将重复的行数进行填充
            print(f'第{i + 1}行数据为重复的数据')
    wb.save('./create_data/014_重复数据的检测.xlsx')


if __name__ == '__main__':
    read_excel()
