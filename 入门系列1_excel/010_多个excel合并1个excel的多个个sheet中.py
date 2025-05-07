from openpyxl import load_workbook, Workbook
import os


# 操作与合并到一个sheet中的操作一样
def copy_data():
    wb = Workbook()
    for name in os.listdir('data'):
        path = f'data/{name}'
        tmp_wb = load_workbook(path)
        tmp_sh = tmp_wb.active
        sh = wb.create_sheet(name.split('.')[0].split("_")[-1])  # 创建一个新的工作簿，对每个excel表都会创建出一个来
        # 先确定行，再确定列，然后按单元格拿取数据（拿到的数据是按每一行来拿取的），然后将每行数据都保存在一个列表里面，再append对应的列表就行
        for r in range(1, tmp_sh.max_row + 1):
            # 获得整行的数据
            row_value = []
            for c in range(1, tmp_sh.max_column + 1):
                value = tmp_sh.cell(r, c).value
                row_value.append(value)  # 往列表里面添加读取到的数据，最后这个列表里面的是一行的数据
            sh.append(row_value)  # 这个保存数据的操作应该要放到读取每一个excel文件后的操作里面
    del wb['Sheet']  # 删除wb里面的名字为“sheet”的工作簿
    # 保存数据
    wb.save('010_合并的数据2.xlsx')


if __name__ == '__main__':
    print(os.listdir('data'))  # 返回到上一级去拿取文件
    copy_data()
    pass
