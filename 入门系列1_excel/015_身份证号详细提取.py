from openpyxl import load_workbook
from datetime import datetime

def create_time():
    wb=load_workbook('身份证信息.xlsx')
    sh=wb.active
    now_year=datetime.now().year
    max_column=sh.max_column
    for i,r in enumerate(sh["B"]):#用枚举来得到对应的行的信息
        id=r.value
        #6位行政区划 4位代表年 2位月 2位日 4位个人识别码
        year=id[6:10]
        mouth=id[10:12]
        day=id[12:14]
        person_id=id[14:18]
        age=now_year-int(year)
        sh.cell(i+1,max_column+1).value=f"{year}年{mouth}月{day}日"#因为索引是从0开始的，所以要加1
        sh.cell(i+1,max_column+2).value=f"年龄为{age}"
        sh.cell(i+1,max_column+3).value=f"个人识别码为{person_id}"
    wb.save("./create_data./015_个人身份信息码.xlsx")


if __name__ == '__main__':
    create_time()