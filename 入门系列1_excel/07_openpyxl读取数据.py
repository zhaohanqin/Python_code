def open():
    from openpyxl import load_workbook
    wb = load_workbook('05_数据汇总案例.xlsx')
    sh1 = wb.active  # 默认激活的工作簿是第一个
    sh2 = wb['案例1']
    # sh3 = wb.get_sheet_by_name('案例1')#这个方法不推荐使用
    # print(sh1 is sh2 is sh3)#证明这三种方法取出来的excel的工作簿是一样的


def show_sheets():
    from openpyxl import load_workbook
    wb = load_workbook('05_数据汇总案例.xlsx')
    print(wb.sheetnames)  # 拿到所有的工作簿的名字，和xlrd有点不同，这里是属性，xlrd里是一个方法

    # 遍历工作簿的名字
    for sh in wb:
        print(sh.title)


# 获取一个数据
def get_one_value():
    from openpyxl import load_workbook
    wb = load_workbook('05_数据汇总案例.xlsx')
    sh = wb['案例1']
    print(sh.cell(2, 3).value)  # 注意，与xlrd不同的是，openpyxl的索引是从1开始的
    print(sh['c2'].value)


# 获取多组数据
def get_many_values():
    from openpyxl import load_workbook
    wb = load_workbook('05_数据汇总案例.xlsx')
    sh = wb.active
    # 切片
    cells1 = sh['c2':'d3']
    # print(cells1)

    # 获取整行整列
    cell_row3 = sh[3]  # 获取第3行的数据
    cell_col3 = sh['c']  # 获取第c列的数据
    print(cell_row3)
    print(cell_col3)
    cell_col3_5 = sh[3:5]
    print(cell_col3_5)  # 返回的是一个元组，里面的元素是每一行的数据，每一行在一个元组里面

    # 通过迭代获取数据
    for row in sh.iter_rows(min_row=2, max_row=5, min_col=3, max_col=3):  # 如果只想拿到第三列的数据的话，就可以将最大最小都设置为3
        for cell in row:
            print(cell.value)

    for row in sh.iter_rows(min_row=2, max_row=5, min_col=2, max_col=3):
        for cell in row:
            print(cell.value)


# 获取所有的数据
def get_all_data():
    from openpyxl import load_workbook
    wb = load_workbook('05_数据汇总案例.xlsx')
    sh = wb.active
    for row in sh.rows:  # 拿到所有的行
        for cell in row:  # 遍历每一行的单元格，拿数据
            print(cell.value)

    for column in sh.columns:  # 拿到所有的列
        for cell in column:  # 遍历每一列的单元格，拿数据
            print(cell.value)


#
def get_num():
    from openpyxl import load_workbook
    wb = load_workbook('05_数据汇总案例.xlsx')
    sh = wb.active
    print(sh.max_row)  # 与xlrd略有不同，xlrd是用sh.nrows
    print(sh.max_column)


if __name__ == '__main__':
    # open()
    # show_sheets()
    # get_one_value()
    # get_many_values()
    # get_all_data()
    get_num()
